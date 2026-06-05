"""Single entry point for ingesting a file.

Routes each file by its inferred Format:

    PROSE       → Unstructured.io + recursive character splitter (existing path)
    CSV         → one chunk per row, pipe-formatted key: value pairs (.csv + .tsv)
    EXCEL       → one chunk per row, pipe-formatted, per sheet (.xlsx via openpyxl,
                  .xls via xlrd) — same self-describing row format as CSV
    everything  → tree-sitter (language-specific grammar) + deterministic NL header

Semantic doc_type is attached to every chunk's metadata for later
retrieval-time filtering and boosting.
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from backend.chunking.chunker import chunk_parsed_document_async
from backend.chunking.code_chunker import chunk_code
from backend.chunking.doc_type import DocType, Format, classify

logger = logging.getLogger(__name__)


async def process_file(
    file_path: str | Path,
    extra_metadata: dict[str, Any] | None = None,
    display_name: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parse + chunk a file. Returns (chunks, doc_level_metadata)."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    logical_path = Path(display_name) if display_name else path

    meta = dict(extra_metadata or {})
    hints = [
        meta.get("source", ""),
        meta.get("category", ""),
        meta.get("folder_path", ""),
    ]
    classification = classify(logical_path, hints=hints)
    meta["format"] = classification.format.value
    meta["doc_type"] = classification.doc_type.value

    logger.info(
        "Processing %s → format=%s doc_type=%s",
        logical_path.name, classification.format.value, classification.doc_type.value,
    )

    if classification.format == Format.PROSE:
        return await _process_prose(path, meta, display_name=logical_path.name)

    if classification.format == Format.CSV:
        # .tsv is tab-separated; everything else on this path is comma-separated.
        delimiter = "\t" if logical_path.suffix.lower() == ".tsv" else ","
        return await asyncio.to_thread(
            _process_csv, path, meta, logical_path.name, delimiter
        )

    if classification.format == Format.EXCEL:
        return await asyncio.to_thread(
            _process_spreadsheet, path, meta, logical_path.name
        )

    return await asyncio.to_thread(
        _process_structured, path, classification.format, meta, logical_path.name
    )


# ── Prose path (existing Unstructured + recursive char splitter) ──────


async def _process_prose(
    path: Path,
    meta: dict[str, Any],
    display_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    from backend.parsing.parser import parse_file
    parsed_doc = await asyncio.to_thread(parse_file, str(path), meta)
    parsed_doc.filename = display_name
    chunks = await chunk_parsed_document_async(parsed_doc)
    doc_meta = {
        "filetype": parsed_doc.filetype,
        "original_filename": display_name,
        "num_elements": len(parsed_doc.elements),
        "format": Format.PROSE.value,
        "doc_type": meta.get("doc_type", DocType.GENERIC.value),
    }
    return chunks, doc_meta


# ── Structured path (tree-sitter) ─────────────────────────────────────


def _process_structured(
    path: Path,
    fmt: Format,
    meta: dict[str, Any],
    display_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    source = path.read_text(encoding="utf-8", errors="replace")
    chunks = chunk_code(
        source=source,
        fmt=fmt,
        filename=display_name,
        metadata=meta,
    )
    doc_meta = {
        "filetype": _mime_for_format(fmt),
        "original_filename": display_name,
        "num_elements": len(chunks),
        "format": fmt.value,
        "doc_type": meta.get("doc_type", DocType.GENERIC.value),
    }
    return chunks, doc_meta


# ── CSV path (direct pipe-format, no tree-sitter) ─────────────────────


def _process_csv(
    path: Path,
    meta: dict[str, Any],
    display_name: str,
    delimiter: str = ",",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parse a CSV/TSV into one chunk per row, each row rendered as
    "col1: value1 | col2: value2 | ...".

    `delimiter` is "," for .csv and "\\t" for .tsv (set by the caller from the
    file extension).

    Rationale: JSON-formatted CSV row chunks were poorly aligned with Nomic's
    NL training distribution (keys and braces drowned out the semantic
    content), which hurt CMDB retrieval recall. Pipe-formatted `key: value`
    pairs read more like natural language and embed closer to NL queries.
    Tree-sitter is no longer involved in the CSV path.

    Defensive steps for real-world CSVs:
      - Strip UTF-8 BOM if present.
      - Use csv.DictReader which handles quoted fields + embedded newlines.
      - Sanitize `\\n` and `\\r` out of every cell value so each row stays a
        single logical line of text.
      - Drop empty cells per row to keep semantic density.
    """
    raw = path.read_text(encoding="utf-8", errors="replace")
    if raw.startswith("\ufeff"):  # strip BOM
        raw = raw[1:]

    rows = _csv_to_row_objects(raw, delimiter=delimiter)
    if not rows:
        return [], _csv_doc_meta(
            display_name, meta, num_elements=0, table_schema={"tables": []}
        )

    chunks: list[dict[str, Any]] = []
    bracket = _bracket_header(display_name, meta)
    for i, row in enumerate(rows):
        pipe_content = _row_to_pipe(row)
        if not pipe_content:
            continue
        chunks.append({
            "text": f"{bracket}{pipe_content}",
            "metadata": {
                **meta,
                "csv_row_index": i,
                "total_rows": len(rows),
                # Structured copy of the row — same data as pipe_content but
                # unambiguous — so the tabular-SQL path can rebuild an exact
                # table without reverse-parsing the pipe string (cell values
                # may legitimately contain '|' or ':'). See _table_schema_entry.
                "row_data": dict(row),
            },
            "element_types": ["csv_row"],
        })

    table_schema = {"tables": [_table_schema_entry(rows, sheet_name=None)]}
    return chunks, _csv_doc_meta(
        display_name, meta, num_elements=len(chunks), table_schema=table_schema
    )


def _csv_to_row_objects(raw: str, delimiter: str = ",") -> list[dict[str, str]]:
    """Parse CSV/TSV into a list of dicts. Sanitizes embedded newlines in cell
    values and drops empty cells. Preserves column order from DictReader."""
    reader = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
    if reader.fieldnames is None:
        return []
    out: list[dict[str, str]] = []
    for row in reader:
        cleaned: dict[str, str] = {}
        for k, v in row.items():
            if k is None or v is None:
                continue
            key = _sanitize_cell(str(k))
            val = _sanitize_cell(str(v))
            if key == "" or val == "":
                continue
            cleaned[key] = val
        if cleaned:
            out.append(cleaned)
    return out


def _sanitize_cell(value: str) -> str:
    """Force a single CSV cell value to a single logical line of text.

    Removes embedded newlines and carriage returns (which otherwise break
    the one-chunk-per-row invariant for the pipe format). Strips surrounding
    whitespace. Does NOT drop pipe characters — if a customer's cell value
    legitimately contains `|`, we accept minor ambiguity in the output string
    rather than destroying the content.
    """
    return value.replace("\n", " ").replace("\r", " ").strip()


def _row_to_pipe(row: dict[str, str]) -> str:
    """Render a row dict as 'col1: val1 | col2: val2 | ...'. Order preserved
    from DictReader insertion order (Python ≥3.7 dicts are ordered)."""
    return " | ".join(f"{k}: {v}" for k, v in row.items())


def _bracket_header(
    display_name: str,
    meta: dict[str, Any],
    sheet_name: str | None = None,
) -> str:
    """Citation-metadata bracket for a tabular row chunk (CSV/TSV/Excel). No
    'Row X of Y' — per production review, the row index carries no semantic
    signal for the LLM and wastes tokens.

    `sheet_name` is included only for multi-sheet workbooks (the caller passes
    None for single-sheet files and CSV/TSV, where a sheet label is noise)."""
    parts = [f"File: {display_name}"]
    if sheet_name:
        parts.append(f"Sheet: {sheet_name}")
    source = meta.get("source", "")
    if source and source not in ("manual", "upload", ""):
        parts.append(f"Source: {source}")
    folder = meta.get("folder_path", "")
    if folder and folder != "/":
        parts.append(f"Path: {folder}")
    return "[" + " | ".join(parts) + "]\n"


def _csv_doc_meta(
    display_name: str,
    meta: dict[str, Any],
    num_elements: int,
    table_schema: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "filetype": "text/csv",
        "original_filename": display_name,
        "num_elements": num_elements,
        "format": Format.CSV.value,
        "doc_type": meta.get("doc_type", DocType.GENERIC.value),
        "table_schema": table_schema or {"tables": []},
    }


# ── Excel path (.xlsx via openpyxl, .xls via xlrd → same pipe format) ──


def _process_spreadsheet(
    path: Path,
    meta: dict[str, Any],
    display_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read an Excel workbook and emit one chunk per data row, rendered in the
    same "col1: value1 | col2: value2 | ..." pipe format as CSV.

    This is the structured-tabular treatment Excel previously MISSED: until
    now .xlsx/.xls fell through to the PROSE/Unstructured path, which rendered
    the whole sheet as one Markdown table and then char-split it. That dropped
    the column-header row after the first chunk, leaving every later chunk as
    headerless columns of numbers the LLM couldn't interpret. Attaching the
    column name to every value in every row (pipe format) makes each row chunk
    self-describing regardless of where retrieval slices.

    Behaviour and known limitations (documented deliberately — real customer
    spreadsheets are messy):
      - Multi-sheet workbooks: every sheet is processed; the sheet name is put
        in the bracket header (only when the workbook has >1 sheet) and in each
        chunk's metadata.
      - HEADER ASSUMPTION: the first non-empty row of each sheet is treated as
        the column header; every subsequent row becomes one chunk. Sheets with
        title banners, merged cells, multiple tables stacked in one sheet, or
        trailing summary rows are NOT specially handled — they degrade to
        best-effort rows. This matches the CSV path's "row 0 is the header"
        assumption and the project's defensive-but-not-magic chunker stance.
      - Empty cells are dropped per row (semantic density), so rows with ragged
        column counts (a real pattern — see the Sample Data.xlsx drift across
        years) still produce labelled values rather than shifted columns.
      - Formula cells: .xlsx is read data_only — the last value Excel cached is
        used; if the file was never opened in Excel that value may be missing.
    """
    suffix = Path(display_name).suffix.lower()
    sheets = _read_workbook(path, suffix)
    multi_sheet = len(sheets) > 1

    chunks: list[dict[str, Any]] = []
    table_entries: list[dict[str, Any]] = []
    for sheet_name, raw_rows in sheets:
        rows = _sheet_rows_to_objects(raw_rows)
        if not rows:
            continue
        # One schema entry per sheet — a multi-sheet workbook is multiple tables
        # with independent columns; the tabular-SQL path builds one DuckDB table
        # per sheet (chunks carry sheet_name to scope reconstruction).
        table_entries.append(_table_schema_entry(rows, sheet_name=sheet_name))
        bracket = _bracket_header(
            display_name, meta, sheet_name=sheet_name if multi_sheet else None
        )
        for i, row in enumerate(rows):
            pipe_content = _row_to_pipe(row)
            if not pipe_content:
                continue
            chunks.append({
                "text": f"{bracket}{pipe_content}",
                "metadata": {
                    **meta,
                    "sheet_name": sheet_name,
                    "csv_row_index": i,
                    "total_rows": len(rows),
                    "row_data": dict(row),
                },
                "element_types": ["spreadsheet_row"],
            })

    table_schema = {"tables": table_entries}
    return chunks, _excel_doc_meta(
        display_name, meta, suffix, num_elements=len(chunks), table_schema=table_schema
    )


def _read_workbook(path: Path, suffix: str) -> list[tuple[str, list[list[Any]]]]:
    """Read every sheet of an Excel workbook as (sheet_name, rows), where each
    row is a list of raw cell values. Engine is chosen by extension, not by the
    temp file's name: openpyxl for .xlsx, xlrd for legacy .xls.
    """
    if suffix == ".xls":
        return _read_xls(path)
    return _read_xlsx(path)


def _read_xlsx(path: Path) -> list[tuple[str, list[list[Any]]]]:
    from openpyxl import load_workbook

    # read_only streams rows (memory-safe for big sheets); data_only returns
    # cached formula results rather than formula strings.
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        out: list[tuple[str, list[list[Any]]]] = []
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            out.append((ws.title, rows))
        return out
    finally:
        wb.close()


def _read_xls(path: Path) -> list[tuple[str, list[list[Any]]]]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    out: list[tuple[str, list[list[Any]]]] = []
    for sh in book.sheets():
        rows: list[list[Any]] = []
        for r in range(sh.nrows):
            row: list[Any] = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                # xls stores dates as floats; convert back to datetime so the
                # formatter renders them like the .xlsx path.
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                        continue
                    except (ValueError, OverflowError):
                        pass
                row.append(cell.value)
            rows.append(row)
        out.append((sh.name, rows))
    return out


def _sheet_rows_to_objects(raw_rows: list[list[Any]]) -> list[dict[str, str]]:
    """Turn a sheet's raw rows into header-keyed dicts, mirroring the CSV path.

    The first non-empty row is the header. Subsequent rows are zipped against
    it; empty cells and unnamed columns are dropped. Blank header cells get a
    synthetic `column_N` name so their data isn't silently lost.
    """
    header_idx = -1
    for i, row in enumerate(raw_rows):
        if any(_format_cell(c) for c in row):
            header_idx = i
            break
    if header_idx == -1:
        return []

    raw_header = raw_rows[header_idx]
    headers = [
        _format_cell(c) or f"column_{j + 1}" for j, c in enumerate(raw_header)
    ]

    out: list[dict[str, str]] = []
    for row in raw_rows[header_idx + 1:]:
        cleaned: dict[str, str] = {}
        for j, col_name in enumerate(headers):
            val = _format_cell(row[j]) if j < len(row) else ""
            if not col_name or not val:
                continue
            cleaned[col_name] = val
        if cleaned:
            out.append(cleaned)
    return out


def _format_cell(value: Any) -> str:
    """Render one spreadsheet cell as a clean, single-line string.

    Handles the value types openpyxl/xlrd hand back: datetimes (trim the
    midnight time component), integral floats (125.0 → "125"), other floats
    (Python's shortest round-trip repr), None/NaN → "". Embedded newlines are
    stripped so each row stays one logical line (the pipe-format invariant)."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value:  # NaN (blank cell via pandas-style readers)
            return ""
        if value.is_integer():
            value = int(value)
        else:
            return _sanitize_cell(str(value))
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):  # plain date (no time component)
        return value.strftime("%Y-%m-%d")
    return _sanitize_cell(str(value))


def _excel_doc_meta(
    display_name: str,
    meta: dict[str, Any],
    suffix: str,
    num_elements: int,
    table_schema: dict[str, Any] | None = None,
) -> dict[str, Any]:
    mime = (
        "application/vnd.ms-excel"
        if suffix == ".xls"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return {
        "filetype": mime,
        "original_filename": display_name,
        "num_elements": num_elements,
        "format": Format.EXCEL.value,
        "doc_type": meta.get("doc_type", DocType.GENERIC.value),
        "table_schema": table_schema or {"tables": []},
    }


# ── Tabular schema inference (shared CSV/Excel) ──────────────────────

_INT_RE = re.compile(r"^[+-]?\d+$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_TIMESTAMP_RE = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")


def _table_schema_entry(
    rows: list[dict[str, str]],
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Derive a column schema for one table (a CSV/TSV, or one Excel sheet).

    Columns are the union of keys across all rows, in first-appearance order —
    rows can differ because empty cells are dropped per row (see
    _csv_to_row_objects / _sheet_rows_to_objects), and real sheets drift their
    columns across rows. Each column's type is inferred from the values that
    actually appear in it. Stored on the document so the tabular-SQL path can
    build a correctly typed DuckDB table without re-reading the original file
    (the original bytes are not retained after ingest).
    """
    col_order: list[str] = []
    col_values: dict[str, list[str]] = {}
    for row in rows:
        for key, val in row.items():
            if key not in col_values:
                col_values[key] = []
                col_order.append(key)
            col_values[key].append(val)
    columns = [
        {"name": c, "type": _infer_column_type(col_values[c])} for c in col_order
    ]
    return {
        "sheet_name": sheet_name,
        "columns": columns,
        "row_count": len(rows),
    }


def _infer_column_type(values: list[str]) -> str:
    """Best-effort SQL-ish type for a column from its string values.

    Conservative: a column is only narrowed to a numeric/temporal type when
    EVERY non-empty value matches. Anything mixed or ambiguous (incl. thousands
    separators like "1,234", or inf/nan) stays 'text' — the SQL builder can
    still TRY_CAST at query time. Returns: integer | double | date |
    timestamp | text.
    """
    vals = [v for v in values if v]
    if not vals:
        return "text"
    if all(_INT_RE.match(v) for v in vals):
        return "integer"
    if all(_is_finite_float(v) for v in vals):
        return "double"
    if all(_DATE_RE.match(v) and _is_valid_date(v, "%Y-%m-%d") for v in vals):
        return "date"
    if all(
        _TIMESTAMP_RE.match(v) and _is_valid_date(v, "%Y-%m-%d %H:%M:%S")
        for v in vals
    ):
        return "timestamp"
    return "text"


def _is_finite_float(value: str) -> bool:
    """True only for a real finite number — rejects 'inf'/'nan'/'infinity'
    (which float() accepts) so text columns aren't misread as numeric."""
    try:
        return math.isfinite(float(value))
    except ValueError:
        return False


def _is_valid_date(value: str, fmt: str) -> bool:
    try:
        datetime.strptime(value, fmt)
        return True
    except ValueError:
        return False


# ── Helpers ───────────────────────────────────────────────────────────


def _mime_for_format(fmt: Format) -> str:
    return {
        Format.TERRAFORM: "text/x-terraform",
        Format.YAML: "application/x-yaml",
        Format.JSON: "application/json",
        Format.PUPPET: "text/x-puppet",
        Format.PYTHON: "text/x-python",
        Format.GO: "text/x-go",
        Format.RUBY: "text/x-ruby",
        Format.JAVASCRIPT: "application/javascript",
        Format.TYPESCRIPT: "application/typescript",
        Format.BASH: "application/x-sh",
        Format.DOCKERFILE: "text/x-dockerfile",
        Format.CSV: "text/csv",
        Format.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        Format.PROSE: "text/plain",
    }.get(fmt, "application/octet-stream")
